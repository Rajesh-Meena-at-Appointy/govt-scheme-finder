import json
import os
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

INPUT_PATH = os.path.join("scripts", "input", "schemes.xlsx")
OUTPUT_PATH = os.path.join("scripts", "output", "schemes.json")

REQUIRED = ["name", "summary", "category", "states", "applyLink"]

def slugify(s: str) -> str:
  s = (s or "").strip().lower()
  out = []
  for ch in s:
    if ch.isalnum():
      out.append(ch)
    elif ch in [" ", "-", "_"]:
      out.append("-")
  slug = "".join(out)
  while "--" in slug:
    slug = slug.replace("--", "-")
  return slug.strip("-")

def split_list(value: Any, sep: str = "|") -> List[str]:
  if value is None:
    return []
  v = str(value).strip()
  if not v:
    return []
  # accept newline or | as separators
  parts = []
  for chunk in v.replace("\r", "\n").split("\n"):
    chunk = chunk.strip()
    if not chunk:
      continue
    parts.extend([p.strip() for p in chunk.split(sep) if p.strip()])
  return parts

def read_rows() -> List[Dict[str, Any]]:
  wb = load_workbook(INPUT_PATH)
  ws = wb.active

  headers = []
  for cell in ws[1]:
    headers.append(str(cell.value).strip() if cell.value is not None else "")

  rows: List[Dict[str, Any]] = []
  for r in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None or str(v).strip() == "" for v in r):
      continue
    obj: Dict[str, Any] = {}
    for i, v in enumerate(r):
      key = headers[i] if i < len(headers) else f"col_{i}"
      if key:
        obj[key] = v
    rows.append(obj)
  return rows

def main():
  if not os.path.exists(INPUT_PATH):
    raise SystemExit(f"Input Excel not found: {INPUT_PATH}")

  os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

  items: List[Dict[str, Any]] = []
  for row in read_rows():
    for req in REQUIRED:
      if req not in row or row[req] is None or str(row[req]).strip() == "":
        raise SystemExit(f"Missing required column/value: {req}. Row: {row}")

    name = str(row.get("name", "")).strip()
    summary = str(row.get("summary", "")).strip()
    category = slugify(str(row.get("category", "other")))
    states_raw = str(row.get("states", "all")).strip().lower()
    states = ["all"] if states_raw == "all" else [slugify(s) for s in states_raw.split(",") if slugify(s)]

    slug = row.get("slug")
    slug = slugify(str(slug)) if slug else slugify(name)

    _id = row.get("id")
    _id = slugify(str(_id)) if _id else slug

    tags = [t.strip() for t in str(row.get("tags", "")).split(",") if t and str(t).strip()]
    benefits = split_list(row.get("benefits", ""), sep="|")
    documents = split_list(row.get("documents", ""), sep="|")

    apply_link = str(row.get("applyLink", "")).strip()

    min_age = int(row.get("minAge", 0) or 0)
    income_max_raw = row.get("incomeMax", None)
    income_max: Optional[int]
    if income_max_raw is None or str(income_max_raw).strip() == "":
      income_max = None
    else:
      income_max = int(float(income_max_raw))

    gender = str(row.get("gender", "any") or "any").strip().lower()
    if gender not in ["any", "male", "female", "other"]:
      gender = "any"

    items.append({
      "id": _id,
      "slug": slug,
      "name": name,
      "summary": summary,
      "category": category,
      "states": states,
      "tags": tags,
      "benefits": benefits,
      "documents": documents,
      "applyLink": apply_link,
      "rules": {
        "minAge": min_age,
        "incomeMax": income_max,
        "gender": gender
      }
    })

  with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(items, f, ensure_ascii=False, indent=2)

  print(f"✅ Generated: {OUTPUT_PATH} (items: {len(items)})")

if __name__ == "__main__":
  main()
